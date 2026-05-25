"""Extract 大阪海外市場調查_dashboard.xlsx into data/osaka-dashboard.json.

Runtime is pure Node.js; this script is a one-shot extractor so the chat
backend never has to parse xlsx at request time.

Each sheet follows the same layout:
  row 0: title
  row 1: source image
  row 2: 數值類型  (percentage or rating)
  row 3: 驗證 / 備註
  row 4: header   (項目 + 12 markets)
  row 5+: data rows; rows whose item is 圖上總計/平均, 計算總計, 差異 are
          validation only and dropped.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / "大阪海外市場調查_dashboard.xlsx"
OUT_PATH = ROOT / "data" / "osaka-dashboard.json"

SKIP_ITEMS = {"圖上總計/平均", "計算總計", "差異", "計算平均"}
SOURCE_SHEET = "目錄"


def format_value(value, value_type: str) -> str:
    if value is None:
        return ""
    if value_type == "百分比":
        return f"{round(float(value) * 100, 1)}%"
    if value_type == "評分":
        return f"{round(float(value), 2)}"
    return str(value)


def extract_sheet(ws) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    title = rows[0][0] if rows else ws.title
    value_type = rows[2][1] if len(rows) > 2 else ""
    note = rows[3][1] if len(rows) > 3 else ""
    header = rows[4] if len(rows) > 4 else ()
    markets = [m for m in header[1:] if m]

    items = []
    data_rows = []
    for raw in rows[5:]:
        item = raw[0]
        if item is None:
            continue
        item = str(item).strip()
        if not item or item in SKIP_ITEMS:
            continue
        values = {}
        for col_idx, market in enumerate(markets, start=1):
            cell = raw[col_idx] if col_idx < len(raw) else None
            if cell is None:
                continue
            try:
                values[market] = float(cell)
            except (TypeError, ValueError):
                continue
        if not values:
            continue
        items.append(item)
        data_rows.append({"item": item, "values": values})

    return {
        "sheet_id": ws.title,
        "title": str(title or ws.title),
        "value_type": str(value_type or ""),
        "note": str(note or ""),
        "markets": markets,
        "items": items,
        "rows": data_rows,
    }


def main() -> None:
    if not XLSX_PATH.exists():
        sys.exit(f"missing xlsx: {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    sheets = []
    cells = []
    for name in wb.sheetnames:
        if name == SOURCE_SHEET:
            continue
        sheet = extract_sheet(wb[name])
        sheets.append(sheet)
        for row in sheet["rows"]:
            for market, value in row["values"].items():
                cells.append(
                    {
                        "sheet_id": sheet["sheet_id"],
                        "sheet_title": sheet["title"],
                        "value_type": sheet["value_type"],
                        "item": row["item"],
                        "market": market,
                        "value": value,
                        "display": format_value(value, sheet["value_type"]),
                    }
                )

    all_markets = []
    for sheet in sheets:
        for m in sheet["markets"]:
            if m not in all_markets:
                all_markets.append(m)

    payload = {
        "source_file": XLSX_PATH.name,
        "markets": all_markets,
        "sheets": sheets,
        "cells": cells,
    }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {OUT_PATH} sheets={len(sheets)} cells={len(cells)}")


if __name__ == "__main__":
    main()
